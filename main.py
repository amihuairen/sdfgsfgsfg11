from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.progressbar import ProgressBar
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.clock import Clock
from kivy.properties import StringProperty
from kivy.storage.jsonstore import JsonStore
from plyer import vibrator
import random
import string
import time
import threading
import openpyxl  # 用于导出Excel


class ActivationScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.store = JsonStore('app_data.json')

        # 如果是全新安装，生成随机激活码
        if not self.store.exists('required_code'):
            random_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            self.store.put('required_code', value=random_code)

        self.required_code = self.store.get('required_code')['value']

        layout = BoxLayout(orientation='vertical', padding=20, spacing=15)

        title = Label(text='抖音评论手机号提取器', font_size=24, bold=True)
        tip = Label(text='这是你的专属激活码\n（仅当前设备有效）', font_size=18)
        code_label = Label(text=f'激活码：{self.required_code}',
                           font_size=28, bold=True, color=(0, 1, 0, 1))

        self.input = TextInput(hint_text='在这里输入上面的激活码', multiline=False, font_size=20)

        activate_btn = Button(text='激活', size_hint=(1, 0.2), font_size=20)
        activate_btn.bind(on_press=self.activate)

        layout.add_widget(title)
        layout.add_widget(tip)
        layout.add_widget(code_label)
        layout.add_widget(self.input)
        layout.add_widget(activate_btn)

        self.add_widget(layout)

    def activate(self, instance):
        entered = self.input.text.strip().upper()
        if entered == self.required_code and not self.store.exists('activated'):
            self.store.put('activated', value=True)
            self.manager.current = 'main'
        else:
            # 可以加个错误提示，这里先简单弹窗
            popup = Popup(title='错误',
                          content=Label(text='激活码错误或已使用！'),
                          size_hint=(0.7, 0.3))
            popup.open()


class MainScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=20, spacing=15)

        title = Label(text='输入抖音视频链接', font_size=22)
        self.link_input = TextInput(hint_text='https://v.douyin.com/xxxxxx', multiline=False, font_size=18)

        start_btn = Button(text='开始扫描', size_hint=(1, 0.15), font_size=20)
        start_btn.bind(on_press=self.start_scan)

        layout.add_widget(title)
        layout.add_widget(self.link_input)
        layout.add_widget(start_btn)
        self.add_widget(layout)

    def start_scan(self, instance):
        link = self.link_input.text.strip()
        if link:
            scan_screen = self.manager.get_screen('scan')
            scan_screen.video_link = link
            self.manager.current = 'scan'
            scan_screen.start_process()


class ScanScreen(Screen):
    log_text = StringProperty('')
    progress_value = 0
    video_link = ''

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=15, spacing=10)

        self.progress_bar = ProgressBar(max=100, size_hint=(1, 0.1))

        scroll = ScrollView()
        self.log_label = Label(text='', valign='top', halign='left', size_hint_y=None, font_size=16)
        self.log_label.bind(texture_size=self.log_label.setter('size'))
        scroll.add_widget(self.log_label)

        self.done_button = Button(text='完成并导出Excel', disabled=True, size_hint=(1, 0.12), font_size=18)
        self.done_button.bind(on_press=self.done)

        layout.add_widget(Label(text='正在模拟扫描评论区...', font_size=20))
        layout.add_widget(self.progress_bar)
        layout.add_widget(scroll)
        layout.add_widget(self.done_button)
        self.add_widget(layout)

        self.phone_set = set()
        self.total_comments = 2000
        self.extracted_count = random.randint(1000, 1500)

    def start_process(self):
        self.log_label.text = ''
        self.progress_bar.value = 0
        self.phone_set.clear()
        threading.Thread(target=self.fake_extract, daemon=True).start()

    def fake_extract(self):
        self.append_log(f"正在解析视频链接: {self.video_link}")
        time.sleep(1.2)
        vibrator.vibrate(0.2)

        self.append_log("模拟调用抖音接口获取评论...")
        steps = ['获取评论批次', '扫描手机号格式', '全局去重处理']

        for i in range(100):
            Clock.schedule_once(lambda dt, val=i + 1: setattr(self.progress_bar, 'value', val), 0)
            step = random.choice(steps)
            self.append_log(f"{step}... ({i + 1}/100)")
            if random.random() < 0.12:
                vibrator.vibrate(0.08)
            time.sleep(random.uniform(0.08, 0.25))

        # 生成假手机号（全局去重）
        while len(self.phone_set) < self.extracted_count:
            fake_phone = f"1{random.randint(3000000000, 8999999999)}"
            self.phone_set.add(fake_phone)

        self.append_log(f"共扫描评论: {self.total_comments} 条")
        self.append_log(f"提取到唯一手机号: {len(self.phone_set)} 个")
        self.append_log("扫描完成！准备导出Excel...")
        vibrator.vibrate(0.6)

        Clock.schedule_once(lambda dt: self.enable_done(), 0)

    def update_progress(self, value):
        self.progress_bar.value = value

    def append_log(self, text):
        Clock.schedule_once(lambda dt: self._append_log(text), 0)

    def _append_log(self, text):
        self.log_label.text += text + '\n'

    def enable_done(self):
        self.done_button.disabled = False

    def done(self, instance):
        # 导出Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Douyin_Phones"
        ws.cell(row=1, column=1).value = "手机号"
        ws.cell(row=1, column=2).value = "提取时间"
        ws.cell(row=1, column=3).value = "来源视频"

        for i, phone in enumerate(self.phone_set, start=2):
            ws.cell(row=i, column=1).value = phone
            ws.cell(row=i, column=2).value = time.strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=i, column=3).value = self.video_link[:50] + "..." if len(
                self.video_link) > 50 else self.video_link

        wb.save('extracted_phones.xlsx')

        popup = Popup(title='导出成功！',
                      content=Label(text='已生成 extracted_phones.xlsx\n可在文件管理器中查看'),
                      size_hint=(0.85, 0.4))
        popup.open()

        self.manager.current = 'main'


class FakeExtractorApp(App):
    def build(self):
        sm = ScreenManager()
        sm.add_widget(ActivationScreen(name='activation'))
        sm.add_widget(MainScreen(name='main'))
        sm.add_widget(ScanScreen(name='scan'))

        store = JsonStore('app_data.json')
        sm.current = 'main' if store.exists('activated') else 'activation'
        return sm


if __name__ == '__main__':
    FakeExtractorApp().run()